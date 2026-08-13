"""Geração das planilhas nos três formatos do contrato: xlsx, csv e json.

O INSTRUCOES lista "implementar só o `xlsx` no download" entre os erros comuns:
os três precisam funcionar.

A planilha sai SEMPRE do `value` persistido, que é o mesmo que o PUT
substitui. É assim que "a correção chega na planilha" sem nenhum caminho
paralelo.

Formato das colunas (README oficial):

- Cartão de ponto: `Data`, depois `Entrada 1`, `Saída 1`, `Entrada 2`,
  `Saída 2`, … com tantos pares quantos o dia com mais batidas exigir.
- Holerite: `Pág.`, `Mês`, `Ano`, depois uma coluna por verba distinta, na
  ordem de primeira aparição.

SEPARAÇÃO DE RESPONSABILIDADES

Este módulo NÃO decide o que é um problema. Ele recebe a avaliação pronta de
`app/services/warnings_service.py` e apenas a pinta. A regra de negócio fica
fora do código do openpyxl para que interface e planilha usem exatamente a
mesma decisão e nunca divirjam.

    value estruturado
        ↓  warnings_service.avaliar()
    LinhaAvaliada[]
        ↓  este módulo
    xlsx com destaque · csv · json

CSV E JSON NÃO CARREGAM DESTAQUE

CSV não suporta formatação, e acrescentar uma coluna de aviso desviaria do
conjunto de colunas que a especificação define. O JSON exporta a transcrição
como está persistida, e é ele que carrega a evidência de incerteza com
fidelidade total — inclusive os `?` que vivem só no `_raw`. Simular formatação
em texto seria pior que não ter.
"""

from __future__ import annotations

import csv
import io
import json
from dataclasses import dataclass, field
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from app.services.warnings_service import (
    AMARELO,
    BORDA_VERMELHA,
    VERMELHO,
    LinhaAvaliada,
    Severidade,
    avaliar,
)

# Estilo exigido literalmente pelo README: cabeçalho em negrito branco sobre
# fundo #173772, nos dois tipos de planilha.
HEADER_BACKGROUND = "173772"
HEADER_FONT_COLOR = "FFFFFF"

FORMATOS_ACEITOS = ("xlsx", "csv", "json")

_MEDIA_TYPES = {
    "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "csv": "text/csv; charset=utf-8",
    "json": "application/json",
}


class FormatoInvalido(Exception):
    """Formato pedido fora de xlsx|csv|json."""


@dataclass
class Planilha:
    """Uma planilha pronta para download."""

    conteudo: bytes
    media_type: str
    filename: str


@dataclass
class Tabela:
    """Representação neutra de planilha: cabeçalho + linhas de strings.

    Existe para que xlsx e csv compartilhem exatamente a mesma montagem de
    colunas. Se divergissem, a planilha baixada mudaria conforme o formato.

    `avaliacoes` acompanha `rows` posição a posição — uma avaliação por linha,
    na mesma ordem. É o único acoplamento entre dado e apresentação, e ele é
    verificado antes de pintar.
    """

    header: List[str]
    rows: List[List[str]]
    avaliacoes: List[LinhaAvaliada] = field(default_factory=list)
    caminhos: List[List[str]] = field(default_factory=list)
    """Caminho de cada célula dentro do `value`, no formato `pages.0.days.3.date_raw`.

    Existe para a interface de revisão: ela precisa saber onde gravar de volta
    a correção que a pessoa digitou numa célula. String vazia quando a célula
    não corresponde a nenhum campo (verba ausente naquela página, par de
    batidas que aquele dia não tem).

    Os exports ignoram este campo. Ele mora aqui, e não num módulo separado,
    porque as colunas da tabela de revisão precisam ser AS MESMAS da planilha —
    é exigência do README. Montar as duas em lugares diferentes garantiria que
    um dia divergissem.
    """

    def severidade_da_linha(self, indice: int) -> Optional[Severidade]:
        if indice >= len(self.avaliacoes):
            return None
        return self.avaliacoes[indice].severidade


# --------------------------------------------------------------------- montagem


def montar_tabela_cartao_ponto(value: Dict[str, Any]) -> Tabela:
    """Uma linha por dia, na ordem do documento."""
    dias: List[Any] = []  # (indice_da_pagina, indice_do_dia, dia)
    for indice_pagina, page in enumerate(value.get("pages", [])):
        for indice_dia, day in enumerate(page.get("days", [])):
            dias.append((indice_pagina, indice_dia, day))

    max_batidas = max((len(d.get("punches", [])) for _p, _d, d in dias), default=0)
    # Colunas vão em pares Entrada/Saída; um número ímpar de batidas ainda
    # ocupa o par inteiro, deixando a última célula vazia.
    pares = (max_batidas + 1) // 2

    header = ["Data"]
    for indice in range(1, pares + 1):
        header.extend([f"Entrada {indice}", f"Saída {indice}"])

    rows: List[List[str]] = []
    caminhos: List[List[str]] = []

    for indice_pagina, indice_dia, dia in dias:
        raiz = f"pages.{indice_pagina}.days.{indice_dia}"

        linha = [str(dia.get("date_raw", ""))]
        caminho = [f"{raiz}.date_raw"]

        for indice_batida, punch in enumerate(dia.get("punches", [])):
            # `time_hhmm` é o valor interpretado, e preserva os `?` de
            # incerteza. `time_raw` continua no JSON para auditoria.
            linha.append(str(punch.get("time_hhmm", "")))
            caminho.append(f"{raiz}.punches.{indice_batida}.time_hhmm")

        faltando = len(header) - len(linha)
        linha.extend([""] * faltando)
        caminho.extend([""] * faltando)

        rows.append(linha)
        caminhos.append(caminho)

    return Tabela(header=header, rows=rows, caminhos=caminhos)


def montar_tabela_holerite(value: Dict[str, Any]) -> Tabela:
    """Uma linha por página; uma coluna por verba distinta.

    Esta é a transposição descrita no README: o documento é uma lista vertical
    de verbas por página, e a planilha é uma matriz larga.
    """
    paginas = value.get("pages", [])

    # União dos labels de `fields`, na ordem de primeira aparição.
    labels: List[str] = []
    vistos = set()
    for pagina in paginas:
        for campo in pagina.get("fields", []):
            label = str(campo.get("label", ""))
            if label not in vistos:
                vistos.add(label)
                labels.append(label)

    header = ["Pág.", "Mês", "Ano"] + labels

    rows: List[List[str]] = []
    caminhos: List[List[str]] = []

    for indice_pagina, pagina in enumerate(paginas):
        raiz = f"pages.{indice_pagina}"

        # Label repetido na mesma página: mantém a primeira ocorrência.
        valores: Dict[str, str] = {}
        posicoes: Dict[str, int] = {}
        for indice_campo, campo in enumerate(pagina.get("fields", [])):
            label = str(campo.get("label", ""))
            if label not in valores:
                valores[label] = str(campo.get("value", ""))
                posicoes[label] = indice_campo

        linha = [
            str(pagina.get("page", "")),
            str(pagina.get("month", "")),
            str(pagina.get("year", "")),
        ]
        # `Pág.` reflete o índice real do PDF e não é corrigível pela pessoa.
        caminho = ["", f"{raiz}.month", f"{raiz}.year"]

        for label in labels:
            linha.append(valores.get(label, ""))
            caminho.append(
                f"{raiz}.fields.{posicoes[label]}.value" if label in posicoes else ""
            )

        rows.append(linha)
        caminhos.append(caminho)

    return Tabela(header=header, rows=rows, caminhos=caminhos)


def montar_tabela(tipo: str, value: Dict[str, Any]) -> Tabela:
    """Monta a tabela e acopla a avaliação de cada linha.

    As duas listas são construídas a partir da mesma estrutura, na mesma ordem
    — uma linha por dia no cartão de ponto, uma por página no holerite. A
    verificação abaixo existe para que uma divergência futura falhe alto, em
    vez de pintar a linha errada em silêncio.
    """
    tabela = (
        montar_tabela_cartao_ponto(value)
        if tipo == "cartao-ponto"
        else montar_tabela_holerite(value)
    )
    tabela.avaliacoes = avaliar(tipo, value)

    if len(tabela.avaliacoes) != len(tabela.rows):
        raise RuntimeError(
            "Avaliação e linhas da planilha estão fora de sincronia: "
            f"{len(tabela.avaliacoes)} avaliações para {len(tabela.rows)} linhas."
        )

    return tabela


# ------------------------------------------------------------------ renderização


def _render_xlsx(tabela: Tabela, titulo: str) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = titulo

    sheet.append(tabela.header)
    fill = PatternFill("solid", fgColor=HEADER_BACKGROUND)
    font = Font(bold=True, color=HEADER_FONT_COLOR)
    for celula in sheet[1]:
        celula.fill = fill
        celula.font = font
        celula.alignment = Alignment(horizontal="center")

    for linha in tabela.rows:
        sheet.append(linha)

    _pintar_destaques(sheet, tabela)
    _ajustar_larguras(sheet, tabela)

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _pintar_destaques(sheet, tabela: Tabela) -> None:
    """Aplica as cores oficiais, linha a linha.

    Este é o único ponto que sabe pintar. Ele não sabe POR QUE a linha está
    marcada — a decisão veio pronta de `warnings_service`.

    Regras do README: amarelo para batidas ímpares, página vazia ou `?` na
    linha; vermelho para data ou mês não sequencial, com borda esquerda
    `#DC3545` na primeira célula. Vermelho ganha quando os dois valem.
    """
    preenchimentos = {
        Severidade.AMARELO: PatternFill("solid", fgColor=AMARELO),
        Severidade.VERMELHO: PatternFill("solid", fgColor=VERMELHO),
    }
    borda_esquerda = Border(left=Side(style="medium", color=BORDA_VERMELHA))

    for indice in range(len(tabela.rows)):
        severidade = tabela.severidade_da_linha(indice)
        if severidade is None:
            continue

        # +2 porque a linha 1 é o cabeçalho e o openpyxl indexa a partir de 1.
        numero_da_linha = indice + 2
        for coluna in range(1, len(tabela.header) + 1):
            sheet.cell(row=numero_da_linha, column=coluna).fill = preenchimentos[
                severidade
            ]

        if severidade is Severidade.VERMELHO:
            sheet.cell(row=numero_da_linha, column=1).border = borda_esquerda


def _ajustar_larguras(sheet, tabela: Tabela) -> None:
    """Largura aproximada por coluna, só para a planilha ser legível."""
    for indice, titulo in enumerate(tabela.header, start=1):
        maior = len(str(titulo))
        for linha in tabela.rows:
            if indice <= len(linha):
                maior = max(maior, len(str(linha[indice - 1])))
        sheet.column_dimensions[sheet.cell(row=1, column=indice).column_letter].width = (
            min(maior + 2, 40)
        )


def _render_csv(tabela: Tabela) -> bytes:
    buffer = io.StringIO()
    writer = csv.writer(buffer, delimiter=";", quoting=csv.QUOTE_MINIMAL)
    writer.writerow(tabela.header)
    writer.writerows(tabela.rows)
    # utf-8-sig: o BOM faz o Excel em português abrir acentuação corretamente.
    # Delimitador ";" pelo mesmo motivo — é o padrão do Excel pt-BR, e os
    # valores monetários usam vírgula decimal.
    return buffer.getvalue().encode("utf-8-sig")


def _render_json(value: Dict[str, Any]) -> bytes:
    return json.dumps(value, ensure_ascii=False, indent=2).encode("utf-8")


# ------------------------------------------------------------------------- API


def gerar_planilha(
    transcricao_id: str,
    tipo: str,
    value: Dict[str, Any],
    formato: str,
) -> Planilha:
    """Gera a planilha no formato pedido.

    `json` devolve a transcrição como está persistida — é a representação mais
    fiel possível, e a única que não perde a distinção entre `_raw` e
    normalizado.
    """
    if formato not in FORMATOS_ACEITOS:
        raise FormatoInvalido(
            f"Formato inválido: {formato!r}. "
            f"Use um de: {', '.join(FORMATOS_ACEITOS)}."
        )

    filename = f"transcricao-{transcricao_id}.{formato}"

    if formato == "json":
        conteudo = _render_json(value)
    else:
        tabela = montar_tabela(tipo, value)
        titulo = "Cartão de ponto" if tipo == "cartao-ponto" else "Holerite"
        conteudo = (
            _render_xlsx(tabela, titulo)
            if formato == "xlsx"
            else _render_csv(tabela)
        )

    return Planilha(
        conteudo=conteudo,
        media_type=_MEDIA_TYPES[formato],
        filename=filename,
    )
