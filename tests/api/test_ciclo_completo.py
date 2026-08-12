"""Ciclo completo: enviar → acompanhar → corrigir → baixar.

Por que este caso: "o ciclo completo funciona" vale 20% da nota, e o critério
explícito da Quick Filler é "a correção chega na planilha?". Este é o teste
que responde essa pergunta de ponta a ponta, nos três formatos.

Usa um pipeline substituído porque na Fase 1 ainda não existe parser real —
todo o resto do caminho é código de produção.
"""

from __future__ import annotations

import csv
import io
import json

from openpyxl import load_workbook

from tests.conftest import upload

VALUE_EXTRAIDO = {
    "pages": [
        {
            "page": 1,
            "days": [
                {
                    "date_raw": "21/05/2019",
                    "punches": [
                        {"kind": "IN", "time_raw": "08:25", "time_hhmm": "08:25"},
                        {"kind": "OUT", "time_raw": "18:25", "time_hhmm": "18:25"},
                    ],
                },
                {"date_raw": "25/05/2019", "punches": []},
            ],
        }
    ]
}


def _pipeline_fake(pdf_path: str, tipo: str):
    return json.loads(json.dumps(VALUE_EXTRAIDO))


def test_ciclo_completo_correcao_chega_nos_tres_formatos(client_factory, pdf_valido):
    client = client_factory(_pipeline_fake)

    # 1. enviar
    criado = client.post("/api/transcricoes", **upload(pdf_valido, "cartao-ponto"))
    assert criado.status_code == 202
    transcricao_id = criado.json()["id"]

    # 2. acompanhar
    corpo = client.get(f"/api/transcricoes/{transcricao_id}").json()
    assert corpo["status"] == "concluido"
    assert corpo["erro"] is None
    assert corpo["value"] == VALUE_EXTRAIDO

    # 3. corrigir — uma pessoa lê o PDF e ajusta um horário
    corrigido = json.loads(json.dumps(VALUE_EXTRAIDO))
    corrigido["pages"][0]["days"][0]["punches"][0]["time_hhmm"] = "07:15"

    resposta_put = client.put(
        f"/api/transcricoes/{transcricao_id}", json={"value": corrigido}
    )
    assert resposta_put.status_code == 200
    assert resposta_put.json()["value"]["pages"][0]["days"][0]["punches"][0][
        "time_hhmm"
    ] == "07:15"

    # a correção precisa estar persistida, não só ecoada
    relido = client.get(f"/api/transcricoes/{transcricao_id}").json()
    assert relido["value"] == corrigido

    # 4. baixar — nos três formatos, todos refletindo a correção
    url = f"/api/transcricoes/{transcricao_id}/planilha"

    resposta_json = client.get(url, params={"formato": "json"})
    assert resposta_json.status_code == 200
    assert json.loads(resposta_json.content) == corrigido

    resposta_csv = client.get(url, params={"formato": "csv"})
    assert resposta_csv.status_code == 200
    linhas = list(
        csv.reader(io.StringIO(resposta_csv.content.decode("utf-8-sig")), delimiter=";")
    )
    assert linhas[0] == ["Data", "Entrada 1", "Saída 1"]
    assert linhas[1] == ["21/05/2019", "07:15", "18:25"]
    # dia sem batida continua sendo uma linha
    assert linhas[2] == ["25/05/2019", "", ""]

    resposta_xlsx = client.get(url, params={"formato": "xlsx"})
    assert resposta_xlsx.status_code == 200
    planilha = load_workbook(io.BytesIO(resposta_xlsx.content))
    aba = planilha.active
    assert [c.value for c in aba[1]] == ["Data", "Entrada 1", "Saída 1"]
    assert [c.value for c in aba[2]] == ["21/05/2019", "07:15", "18:25"]


def test_cabecalho_da_planilha_usa_o_estilo_oficial(client_factory, pdf_valido):
    """Cabeçalho em negrito branco sobre #173772, exigido literalmente."""
    client = client_factory(_pipeline_fake)
    criado = client.post("/api/transcricoes", **upload(pdf_valido, "cartao-ponto"))
    transcricao_id = criado.json()["id"]

    resposta = client.get(
        f"/api/transcricoes/{transcricao_id}/planilha", params={"formato": "xlsx"}
    )
    aba = load_workbook(io.BytesIO(resposta.content)).active

    for celula in aba[1]:
        assert celula.font.bold is True
        assert celula.font.color.rgb.endswith("FFFFFF")
        assert celula.fill.fgColor.rgb.endswith("173772")


def test_formato_invalido_e_recusado(client_factory, pdf_valido):
    client = client_factory(_pipeline_fake)
    criado = client.post("/api/transcricoes", **upload(pdf_valido, "cartao-ponto"))
    transcricao_id = criado.json()["id"]

    resposta = client.get(
        f"/api/transcricoes/{transcricao_id}/planilha", params={"formato": "pdf"}
    )
    assert resposta.status_code == 400


def test_planilha_do_holerite_transpoe_verbas_em_colunas(client_factory, pdf_valido):
    """A transposição é o trabalho descrito no README para o holerite.

    Verifica também a separação `fields` / `bases`: nenhuma base pode virar
    coluna, porque isso "contamina a planilha inteira".
    """

    value = {
        "pages": [
            {
                "page": 1,
                "year": "2020",
                "month": "01",
                "fields": [
                    {
                        "code": "0010",
                        "label": "Salário Base",
                        "reference": "220,00",
                        "value": "2.389,77",
                    },
                    {
                        "code": "0998",
                        "label": "INSS",
                        "reference": "",
                        "value": "262,87",
                    },
                ],
                "bases": [{"label": "Base INSS", "value": "2.545,68"}],
            },
            {
                "page": 2,
                "year": "2020",
                "month": "02",
                "fields": [
                    {
                        "code": "5560",
                        "label": "Horas Extras - 50%",
                        "reference": "8,00",
                        "value": "155,91",
                    }
                ],
                "bases": [],
            },
        ]
    }

    client = client_factory(lambda pdf_path, tipo: json.loads(json.dumps(value)))
    criado = client.post("/api/transcricoes", **upload(pdf_valido, "holerite"))
    transcricao_id = criado.json()["id"]

    resposta = client.get(
        f"/api/transcricoes/{transcricao_id}/planilha", params={"formato": "csv"}
    )
    linhas = list(
        csv.reader(io.StringIO(resposta.content.decode("utf-8-sig")), delimiter=";")
    )

    # colunas fixas + união dos labels de `fields`, na ordem de aparição
    assert linhas[0] == [
        "Pág.",
        "Mês",
        "Ano",
        "Salário Base",
        "INSS",
        "Horas Extras - 50%",
    ]
    assert "Base INSS" not in linhas[0]

    assert linhas[1] == ["1", "01", "2020", "2.389,77", "262,87", ""]
    # verba ausente na página vira célula vazia
    assert linhas[2] == ["2", "02", "2020", "", "", "155,91"]


def test_dinheiro_permanece_string_no_json(client_factory, pdf_valido):
    """Converter para float perde formato e introduz erro de arredondamento."""
    value = {
        "pages": [
            {
                "page": 1,
                "year": "2020",
                "month": "01",
                "fields": [
                    {
                        "code": "",
                        "label": "Salário Base",
                        "reference": "",
                        "value": "2.389,77",
                    }
                ],
                "bases": [],
            }
        ]
    }
    client = client_factory(lambda pdf_path, tipo: json.loads(json.dumps(value)))
    criado = client.post("/api/transcricoes", **upload(pdf_valido, "holerite"))
    transcricao_id = criado.json()["id"]

    resposta = client.get(
        f"/api/transcricoes/{transcricao_id}/planilha", params={"formato": "json"}
    )
    devolvido = json.loads(resposta.content)

    assert devolvido["pages"][0]["fields"][0]["value"] == "2.389,77"
    assert isinstance(devolvido["pages"][0]["fields"][0]["value"], str)
