"""Destaques no XLSX — inspecionados célula a célula.

Por que estes casos: o arquivo ser gerado sem exceção não prova nada. As cores
são literais da especificação, e uma linha pintada errada é pior que uma linha
sem pintura — leva a pessoa a conferir o registro errado.
"""

from __future__ import annotations

import io
from pathlib import Path

from openpyxl import load_workbook

from app.extraction.native_text import extract_native_pages
from app.parsers.timesheet.sipon import SiponTimesheetParser
from app.services.export_service import gerar_planilha

EXEMPLOS = Path(__file__).resolve().parents[2] / "exemplos"

# O openpyxl devolve as cores com o canal alfa à frente.
AMARELO = "00FFF3CD"
VERMELHO = "00F8D7DA"
BORDA = "00DC3545"
SEM_PREENCHIMENTO = "00000000"


def _abrir(conteudo: bytes):
    return load_workbook(io.BytesIO(conteudo)).active


def _cartao(dias):
    return {"pages": [{"page": 1, "days": dias}]}


def _dia(date_raw, horarios=(), raws=None):
    brutos = raws if raws is not None else horarios
    return {
        "date_raw": date_raw,
        "punches": [
            {"kind": "IN", "time_raw": bruto, "time_hhmm": normal}
            for bruto, normal in zip(brutos, horarios)
        ],
    }


def test_cabecalho_usa_o_estilo_oficial():
    planilha = gerar_planilha(
        "x", "cartao-ponto", _cartao([_dia("01/05/2019", ["08:00", "12:00"])]), "xlsx"
    )
    aba = _abrir(planilha.conteudo)

    for celula in aba[1]:
        assert celula.font.bold is True
        assert celula.font.color.rgb == "00FFFFFF"
        assert celula.fill.fgColor.rgb == "00173772"


def test_linha_sem_problema_nao_recebe_preenchimento():
    planilha = gerar_planilha(
        "x", "cartao-ponto", _cartao([_dia("01/05/2019", ["08:00", "12:00"])]), "xlsx"
    )
    aba = _abrir(planilha.conteudo)

    assert aba.cell(row=2, column=1).fill.fgColor.rgb == SEM_PREENCHIMENTO


def test_batidas_impares_pintam_a_linha_inteira_de_amarelo():
    planilha = gerar_planilha(
        "x", "cartao-ponto", _cartao([_dia("01/05/2019", ["08:00"])]), "xlsx"
    )
    aba = _abrir(planilha.conteudo)

    for coluna in range(1, aba.max_column + 1):
        assert aba.cell(row=2, column=coluna).fill.fgColor.rgb == AMARELO

    # amarelo não leva borda
    assert aba.cell(row=2, column=1).border.left.style is None


def test_data_nao_sequencial_pinta_de_vermelho_com_borda_esquerda():
    valor = _cartao(
        [
            _dia("01/05/2019", ["08:00", "12:00"]),
            _dia("09/05/2019", ["08:00", "12:00"]),
        ]
    )
    aba = _abrir(gerar_planilha("x", "cartao-ponto", valor, "xlsx").conteudo)

    assert aba.cell(row=2, column=1).fill.fgColor.rgb == SEM_PREENCHIMENTO

    for coluna in range(1, aba.max_column + 1):
        assert aba.cell(row=3, column=coluna).fill.fgColor.rgb == VERMELHO

    borda = aba.cell(row=3, column=1).border.left
    assert borda.style is not None
    assert borda.color.rgb == BORDA
    # a borda é só na primeira célula
    assert aba.cell(row=3, column=2).border.left.style is None


def test_vermelho_ganha_de_amarelo_na_planilha():
    """Linha com batida ímpar E data fora de sequência sai vermelha."""
    valor = _cartao([_dia("01/05/2019", ["08:00", "12:00"]), _dia("09/05/2019", ["08:00"])])
    aba = _abrir(gerar_planilha("x", "cartao-ponto", valor, "xlsx").conteudo)

    assert aba.cell(row=3, column=1).fill.fgColor.rgb == VERMELHO
    assert aba.cell(row=3, column=1).border.left.color.rgb == BORDA


def test_interrogacao_apenas_no_raw_pinta_a_linha():
    """A célula mostra `23:00`, sem `?` — e mesmo assim precisa ser destacada.

    É o caso real de `time-card-03`, e a razão de o destaque ser derivado do
    dado estruturado e não do texto da célula.
    """
    valor = _cartao(
        [_dia("27/01/2020", horarios=["14:56", "23:00"], raws=["14:56", "23:00?"])]
    )
    aba = _abrir(gerar_planilha("x", "cartao-ponto", valor, "xlsx").conteudo)

    assert aba.cell(row=2, column=3).value == "23:00"  # sem `?` na célula
    assert aba.cell(row=2, column=1).fill.fgColor.rgb == AMARELO


def test_documento_real_destaca_apenas_a_anomalia_conhecida():
    """`time-card-01`: 153 dias, uma única linha problemática.

    `29/10/2012` tem uma batida só no documento. Nenhuma das outras 152 linhas
    pode ser pintada — falso alarme em massa faria a pessoa ignorar os alertas.
    """
    value = SiponTimesheetParser().parse(
        extract_native_pages(str(EXEMPLOS / "time-card-01.pdf"))
    )
    aba = _abrir(gerar_planilha("x", "cartao-ponto", value, "xlsx").conteudo)

    pintadas = [
        (linha, aba.cell(row=linha, column=1).value)
        for linha in range(2, aba.max_row + 1)
        if aba.cell(row=linha, column=1).fill.fgColor.rgb != SEM_PREENCHIMENTO
    ]

    assert len(pintadas) == 1
    assert pintadas[0][1] == "29/10/2012"
    assert aba.cell(row=pintadas[0][0], column=1).fill.fgColor.rgb == AMARELO


def test_holerite_real_nao_gera_nenhum_destaque():
    """`payroll-03` vai de 10/2019 a 02/2020, atravessando a virada de ano.

    Se dezembro→janeiro fosse tratado como quebra, esta planilha teria uma
    linha vermelha indevida.
    """
    from app.parsers.payslip.demonstrativo_mensal import DemonstrativoMensalParser

    value = DemonstrativoMensalParser().parse(
        extract_native_pages(str(EXEMPLOS / "payroll-03.pdf"))
    )
    aba = _abrir(gerar_planilha("x", "holerite", value, "xlsx").conteudo)

    pintadas = [
        linha
        for linha in range(2, aba.max_row + 1)
        if aba.cell(row=linha, column=1).fill.fgColor.rgb != SEM_PREENCHIMENTO
    ]

    assert pintadas == []


def test_csv_e_json_continuam_sem_coluna_extra():
    """Aviso é apresentação; o conjunto de colunas é contrato.

    Acrescentar uma coluna de aviso desviaria da especificação. O JSON continua
    carregando a evidência de incerteza com fidelidade total.
    """
    import csv as csv_mod
    import json as json_mod

    valor = _cartao([_dia("01/05/2019", ["08:00"])])

    csv_bytes = gerar_planilha("x", "cartao-ponto", valor, "csv").conteudo
    linhas = list(
        csv_mod.reader(io.StringIO(csv_bytes.decode("utf-8-sig")), delimiter=";")
    )
    assert linhas[0] == ["Data", "Entrada 1", "Saída 1"]

    json_bytes = gerar_planilha("x", "cartao-ponto", valor, "json").conteudo
    assert json_mod.loads(json_bytes) == valor
